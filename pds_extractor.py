import pandas as pd
import openpyxl
from openpyxl import load_workbook
import re
from datetime import datetime
import json

class PDSExtractor:
    def __init__(self):
        self.pds_data = {}
        self.errors = []
        self.warnings = []
    
    def extract_pds_data(self, file_path):
        """Main extraction function for PDS files"""
        try:
            wb = load_workbook(file_path, data_only=True)
            
            # Verify this is a PDS file
            if not self._is_pds_file(wb):
                raise ValueError("File does not appear to be a valid CSC PDS format")
            
            # Extract from each sheet
            if 'C1' in wb.sheetnames:
                self.pds_data['personal_info'] = self._extract_c1_personal_info(wb['C1'])
            
            if 'C2' in wb.sheetnames:
                c2_data = self._extract_c2_eligibility_work(wb['C2'])
                self.pds_data['eligibility'] = c2_data.get('eligibility', [])
                self.pds_data['work_experience'] = c2_data.get('work_experience', [])
            
            if 'C3' in wb.sheetnames:
                c3_data = self._extract_c3_voluntary_training(wb['C3'])
                self.pds_data['voluntary_work'] = c3_data.get('voluntary_work', [])
                self.pds_data['training'] = c3_data.get('training', [])
            
            if 'C4' in wb.sheetnames:
                self.pds_data['other_info'] = self._extract_c4_other_info(wb['C4'])
            
            wb.close()
            
            # Add metadata
            self.pds_data['extraction_metadata'] = {
                'extracted_at': datetime.now().isoformat(),
                'file_type': 'CSC_PDS',
                'sheets_processed': list(wb.sheetnames),
                'errors': self.errors,
                'warnings': self.warnings
            }
            
            return self.pds_data
            
        except Exception as e:
            self.errors.append(f"Error extracting PDS data: {str(e)}")
            return None
    
    def _is_pds_file(self, workbook):
        """Check if this is a valid PDS file"""
        sheet_names = workbook.sheetnames
        
        # Check for C1-C4 sheets or variations
        pds_sheets = ['C1', 'C2', 'C3', 'C4']
        has_pds_sheets = any(sheet in sheet_names for sheet in pds_sheets)
        
        # Check for PDS indicators in first sheet
        first_sheet = workbook[sheet_names[0]]
        try:
            # Look for PDS text indicators
            for row in range(1, 10):
                for col in range(1, 10):
                    cell_value = first_sheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str):
                        if 'PERSONAL DATA SHEET' in cell_value.upper():
                            return True
                        if 'CS FORM NO. 212' in cell_value.upper():
                            return True
        except:
            pass
        
        return has_pds_sheets
    
    def _extract_c1_personal_info(self, worksheet):
        """Extract personal information from C1 sheet"""
        personal_info = {}
        
        try:
            # Extract basic personal information
            personal_info['surname'] = self._get_cell_value_by_pattern(worksheet, 'SURNAME', adjacent=True)
            personal_info['first_name'] = self._get_cell_value_by_pattern(worksheet, 'FIRST NAME', adjacent=True)
            personal_info['middle_name'] = self._get_cell_value_by_pattern(worksheet, 'MIDDLE NAME', adjacent=True)
            personal_info['name_extension'] = self._get_cell_value_by_pattern(worksheet, 'NAME EXTENSION', adjacent=True)
            
            # Date and place of birth
            personal_info['date_of_birth'] = self._get_cell_value_by_pattern(worksheet, 'DATE OF BIRTH', adjacent=True)
            personal_info['place_of_birth'] = self._get_cell_value_by_pattern(worksheet, 'PLACE OF BIRTH', adjacent=True)
            
            # Basic demographics
            personal_info['sex'] = self._get_cell_value_by_pattern(worksheet, 'SEX', adjacent=True)
            personal_info['civil_status'] = self._get_cell_value_by_pattern(worksheet, 'CIVIL STATUS', adjacent=True)
            personal_info['height'] = self._get_cell_value_by_pattern(worksheet, 'HEIGHT', adjacent=True)
            personal_info['weight'] = self._get_cell_value_by_pattern(worksheet, 'WEIGHT', adjacent=True)
            personal_info['blood_type'] = self._get_cell_value_by_pattern(worksheet, 'BLOOD TYPE', adjacent=True)
            
            # Government IDs
            personal_info['gsis_id'] = self._get_cell_value_by_pattern(worksheet, 'GSIS ID NO', adjacent=True)
            personal_info['pagibig_id'] = self._get_cell_value_by_pattern(worksheet, 'PAG-IBIG ID NO', adjacent=True)
            personal_info['philhealth_no'] = self._get_cell_value_by_pattern(worksheet, 'PHILHEALTH NO', adjacent=True)
            personal_info['sss_no'] = self._get_cell_value_by_pattern(worksheet, 'SSS NO', adjacent=True)
            personal_info['tin_no'] = self._get_cell_value_by_pattern(worksheet, 'TIN NO', adjacent=True)
            
            # Citizenship
            personal_info['citizenship'] = self._get_cell_value_by_pattern(worksheet, 'CITIZENSHIP', adjacent=True)
            personal_info['dual_citizenship_country'] = self._get_cell_value_by_pattern(worksheet, 'country:', adjacent=True)
            
            # Contact information
            personal_info['residential_address'] = self._extract_address(worksheet, 'RESIDENTIAL ADDRESS')
            personal_info['permanent_address'] = self._extract_address(worksheet, 'PERMANENT ADDRESS')
            personal_info['telephone_no'] = self._get_cell_value_by_pattern(worksheet, 'TELEPHONE NO', adjacent=True)
            personal_info['mobile_no'] = self._get_cell_value_by_pattern(worksheet, 'MOBILE NO', adjacent=True)
            personal_info['email'] = self._get_cell_value_by_pattern(worksheet, 'E-MAIL ADDRESS', adjacent=True)
            
            # Educational background
            personal_info['education'] = self._extract_education(worksheet)
            
            # Family background
            personal_info['family'] = self._extract_family_background(worksheet)
            
        except Exception as e:
            self.errors.append(f"Error extracting C1 personal info: {str(e)}")
        
        return personal_info
    
    def _extract_c2_eligibility_work(self, worksheet):
        """Extract civil service eligibility and work experience from C2"""
        c2_data = {}
        
        try:
            # Extract civil service eligibility
            c2_data['eligibility'] = self._extract_eligibility(worksheet)
            
            # Extract work experience
            c2_data['work_experience'] = self._extract_work_experience(worksheet)
            
        except Exception as e:
            self.errors.append(f"Error extracting C2 data: {str(e)}")
        
        return c2_data
    
    def _extract_c3_voluntary_training(self, worksheet):
        """Extract voluntary work and training from C3"""
        c3_data = {}
        
        try:
            # Extract voluntary work
            c3_data['voluntary_work'] = self._extract_voluntary_work(worksheet)
            
            # Extract training programs
            c3_data['training'] = self._extract_training_programs(worksheet)
            
        except Exception as e:
            self.errors.append(f"Error extracting C3 data: {str(e)}")
        
        return c3_data
    
    def _extract_c4_other_info(self, worksheet):
        """Extract other information and references from C4"""
        other_info = {}
        
        try:
            # Extract questions about relationships, charges, etc.
            other_info['government_relationship'] = self._extract_yes_no_questions(worksheet)
            
            # Extract references
            other_info['references'] = self._extract_references(worksheet)
            
            # Extract government service record
            other_info['government_service'] = self._extract_government_service(worksheet)
            
        except Exception as e:
            self.errors.append(f"Error extracting C4 data: {str(e)}")
        
        return other_info
    
    def _get_cell_value_by_pattern(self, worksheet, pattern, adjacent=False, search_area=(1, 1, 100, 20)):
        """Find a cell containing the pattern and optionally return adjacent cell value"""
        try:
            start_row, start_col, max_row, max_col = search_area
            
            for row in range(start_row, min(max_row + 1, worksheet.max_row + 1)):
                for col in range(start_col, min(max_col + 1, worksheet.max_column + 1)):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str):
                        if pattern.upper() in cell_value.upper():
                            if adjacent:
                                # Try adjacent cells (right, below, two cells right)
                                for offset in [(0, 1), (0, 2), (1, 0), (0, 3)]:
                                    adj_row, adj_col = row + offset[0], col + offset[1]
                                    if adj_row <= worksheet.max_row and adj_col <= worksheet.max_column:
                                        adj_value = worksheet.cell(row=adj_row, column=adj_col).value
                                        if adj_value and str(adj_value).strip():
                                            return str(adj_value).strip()
                            else:
                                return str(cell_value).strip()
        except Exception as e:
            self.warnings.append(f"Error finding pattern '{pattern}': {str(e)}")
        
        return None
    
    def _extract_address(self, worksheet, address_type):
        """Extract address information"""
        address = {}
        try:
            # Look for address components
            if 'RESIDENTIAL' in address_type.upper():
                base_pattern = 'RESIDENTIAL ADDRESS'
            else:
                base_pattern = 'PERMANENT ADDRESS'
            
            # Find the starting position
            for row in range(1, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str) and base_pattern in cell_value.upper():
                        # Extract address components from surrounding cells
                        address['full_address'] = self._collect_address_parts(worksheet, row, col)
                        break
        except Exception as e:
            self.warnings.append(f"Error extracting {address_type}: {str(e)}")
        
        return address
    
    def _collect_address_parts(self, worksheet, start_row, start_col):
        """Collect address parts from multiple cells"""
        address_parts = []
        
        # Look in surrounding area for address components
        for row_offset in range(0, 8):
            for col_offset in range(0, 6):
                try:
                    cell = worksheet.cell(row=start_row + row_offset, column=start_col + col_offset)
                    if cell.value and isinstance(cell.value, str):
                        value = str(cell.value).strip()
                        # Skip labels and empty values
                        if (len(value) > 2 and 
                            not any(label in value.upper() for label in 
                                   ['ADDRESS', 'HOUSE', 'STREET', 'BARANGAY', 'CITY', 'PROVINCE', 'ZIP'])):
                            address_parts.append(value)
                except:
                    continue
        
        return ', '.join(address_parts) if address_parts else None
    
    def _extract_education(self, worksheet):
        """Extract educational background"""
        education = {}
        
        education_levels = ['ELEMENTARY', 'SECONDARY', 'VOCATIONAL', 'COLLEGE', 'GRADUATE']
        
        for level in education_levels:
            education[level.lower()] = self._get_cell_value_by_pattern(worksheet, level, adjacent=True)
        
        return education
    
    def _extract_family_background(self, worksheet):
        """Extract family background information"""
        family = {}
        
        try:
            # Look for family section
            family_patterns = ['SPOUSE', 'FATHER', 'MOTHER', 'CHILDREN']
            
            for pattern in family_patterns:
                family[pattern.lower()] = self._get_cell_value_by_pattern(worksheet, pattern, adjacent=True)
        
        except Exception as e:
            self.warnings.append(f"Error extracting family background: {str(e)}")
        
        return family
    
    def _extract_eligibility(self, worksheet):
        """Extract civil service eligibility data"""
        eligibility_list = []
        
        try:
            # Look for eligibility section starting point
            for row in range(1, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and 'CIVIL SERVICE ELIGIBILITY' in str(cell_value).upper():
                        # Extract eligibility entries from rows below
                        eligibility_list = self._extract_table_data(worksheet, row + 2, 
                                                                  ['eligibility', 'rating', 'date_exam', 'place_exam', 'license_no', 'validity'])
                        break
                if eligibility_list:
                    break
        
        except Exception as e:
            self.warnings.append(f"Error extracting eligibility: {str(e)}")
        
        return eligibility_list
    
    def _extract_work_experience(self, worksheet):
        """Extract work experience data"""
        work_experience = []
        
        try:
            # Look for work experience section
            for row in range(1, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and 'WORK EXPERIENCE' in str(cell_value).upper():
                        # Extract work entries from rows below
                        work_experience = self._extract_table_data(worksheet, row + 3,
                                                                 ['date_from', 'date_to', 'position', 'company', 'salary', 'grade', 'status', 'govt_service'])
                        break
                if work_experience:
                    break
        
        except Exception as e:
            self.warnings.append(f"Error extracting work experience: {str(e)}")
        
        return work_experience
    
    def _extract_voluntary_work(self, worksheet):
        """Extract voluntary work data"""
        voluntary_work = []
        
        try:
            # Look for voluntary work section
            for row in range(1, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and 'VOLUNTARY WORK' in str(cell_value).upper():
                        voluntary_work = self._extract_table_data(worksheet, row + 3,
                                                                ['organization', 'date_from', 'date_to', 'hours', 'position'])
                        break
                if voluntary_work:
                    break
        
        except Exception as e:
            self.warnings.append(f"Error extracting voluntary work: {str(e)}")
        
        return voluntary_work
    
    def _extract_training_programs(self, worksheet):
        """Extract training and development programs"""
        training = []
        
        try:
            # Look for L&D section
            for row in range(1, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and 'LEARNING AND DEVELOPMENT' in str(cell_value).upper():
                        training = self._extract_table_data(worksheet, row + 3,
                                                          ['title', 'date_from', 'date_to', 'hours', 'type', 'conductor'])
                        break
                if training:
                    break
        
        except Exception as e:
            self.warnings.append(f"Error extracting training: {str(e)}")
        
        return training
    
    def _extract_yes_no_questions(self, worksheet):
        """Extract yes/no questions and answers from C4"""
        questions = {}
        
        question_patterns = [
            'related by consanguinity',
            'found guilty of any administrative offense',
            'criminally charged',
            'convicted of any crime',
            'separated from the service'
        ]
        
        for pattern in question_patterns:
            answer = self._find_yes_no_answer(worksheet, pattern)
            questions[pattern.replace(' ', '_')] = answer
        
        return questions
    
    def _find_yes_no_answer(self, worksheet, question_pattern):
        """Find yes/no answer for a specific question"""
        try:
            # Find the question first
            for row in range(1, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and question_pattern.upper() in str(cell_value).upper():
                        # Look for Yes/No in surrounding cells
                        for r_offset in range(-2, 3):
                            for c_offset in range(-2, 8):
                                try:
                                    check_cell = worksheet.cell(row=row + r_offset, column=col + c_offset)
                                    if check_cell.value:
                                        value = str(check_cell.value).upper().strip()
                                        if value in ['YES', 'NO', 'Y', 'N']:
                                            return value
                                except:
                                    continue
        except Exception as e:
            self.warnings.append(f"Error finding answer for '{question_pattern}': {str(e)}")
        
        return None
    
    def _extract_references(self, worksheet):
        """Extract character references"""
        references = []
        
        try:
            # Look for references section (usually at the bottom of C4)
            ref_keywords = ['REFERENCE', 'CHARACTER REFERENCE', 'REFERENCES']
            
            for keyword in ref_keywords:
                for row in range(1, worksheet.max_row + 1):
                    for col in range(1, worksheet.max_column + 1):
                        cell_value = worksheet.cell(row=row, column=col).value
                        if cell_value and keyword in str(cell_value).upper():
                            references = self._extract_table_data(worksheet, row + 2,
                                                                ['name', 'address', 'telephone_no'])
                            if references:
                                return references
        
        except Exception as e:
            self.warnings.append(f"Error extracting references: {str(e)}")
        
        return references
    
    def _extract_government_service(self, worksheet):
        """Extract government service information"""
        # This is usually indicated in work experience or separate section
        return self._get_cell_value_by_pattern(worksheet, 'GOVERNMENT SERVICE', adjacent=True)
    
    def _extract_table_data(self, worksheet, start_row, columns):
        """Extract tabular data starting from a specific row"""
        table_data = []
        
        try:
            # Find data rows (skip empty rows)
            for row in range(start_row, min(start_row + 20, worksheet.max_row + 1)):
                row_data = {}
                has_data = False
                
                for col_idx, col_name in enumerate(columns):
                    cell_value = worksheet.cell(row=row, column=col_idx + 1).value
                    if cell_value and str(cell_value).strip():
                        row_data[col_name] = str(cell_value).strip()
                        has_data = True
                    else:
                        row_data[col_name] = None
                
                if has_data:
                    table_data.append(row_data)
                elif len(table_data) > 0:
                    # Stop if we hit empty rows after finding data
                    break
        
        except Exception as e:
            self.warnings.append(f"Error extracting table data: {str(e)}")
        
        return table_data

# Test function
def test_pds_extraction():
    """Test the PDS extraction with the sample file"""
    extractor = PDSExtractor()
    
    try:
        result = extractor.extract_pds_data("Sample PDS Lenar.xlsx")
        
        if result:
            print("=== PDS EXTRACTION SUCCESSFUL ===")
            print(f"Sections extracted: {list(result.keys())}")
            
            # Show personal info
            if 'personal_info' in result:
                personal = result['personal_info']
                print(f"\nPersonal Info:")
                print(f"  Name: {personal.get('first_name', '')} {personal.get('middle_name', '')} {personal.get('surname', '')}")
                print(f"  Email: {personal.get('email', 'N/A')}")
                print(f"  Birth Date: {personal.get('date_of_birth', 'N/A')}")
            
            # Show work experience count
            if 'work_experience' in result:
                print(f"\nWork Experience: {len(result['work_experience'])} entries")
            
            # Show eligibility count
            if 'eligibility' in result:
                print(f"Eligibility: {len(result['eligibility'])} entries")
            
            # Show training count
            if 'training' in result:
                print(f"Training: {len(result['training'])} entries")
            
            # Save extracted data
            with open('extracted_pds_data.json', 'w', encoding='utf-8') as f:
                json.dump(result, f, indent=2, ensure_ascii=False)
            
            print(f"\nFull extracted data saved to extracted_pds_data.json")
            
            # Show any errors or warnings
            if extractor.errors:
                print(f"\nErrors: {len(extractor.errors)}")
                for error in extractor.errors:
                    print(f"  - {error}")
            
            if extractor.warnings:
                print(f"\nWarnings: {len(extractor.warnings)}")
                for warning in extractor.warnings:
                    print(f"  - {warning}")
            
            return True
        else:
            print("Failed to extract PDS data")
            return False
            
    except Exception as e:
        print(f"Error testing PDS extraction: {e}")
        return False

if __name__ == "__main__":
    test_pds_extraction()